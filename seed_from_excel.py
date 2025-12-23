import os
import math
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from app.db import Base, engine, SessionLocal
from app.models import Ingredient

EXCEL_SHEET = "원재료_DB"

COLMAP = {
    "원재료 선택명(자동: 원재료|브랜드)": "display_name",
    "원재료명": "name",
    "브랜드/제조사": "brand",
    "기준량(g)": "base_g",
    "나트륨(mg/100g)": "sodium_mg_100g",
    "탄수화물(g/100g)": "carbs_g_100g",
    "당류(g/100g)": "sugars_g_100g",
    "식이섬유(g/100g)": "fiber_g_100g",
    "알룰로스(g/100g)": "allulose_g_100g",
    "지방(g/100g)": "fat_g_100g",
    "트랜스지방(g/100g)": "trans_fat_g_100g",
    "포화지방(g/100g)": "sat_fat_g_100g",
    "콜레스테롤(mg/100g)": "chol_mg_100g",
    "단백질(g/100g)": "protein_g_100g",
    "메모(출처/라벨)": "memo",
}

def safe_num(v):
    try:
        if v is None:
            return 0.0
        if isinstance(v, float) and math.isnan(v):
            return 0.0
        return float(v)
    except Exception:
        return 0.0

def main():
    excel_path = os.getenv("EXCEL_PATH", "영양성분계산기_오터.xlsx")
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {excel_path}")


    Base.metadata.create_all(bind=engine)

    wb = load_workbook(excel_path, data_only=True)
    ws = wb[EXCEL_SHEET]

    headers = [cell.value for cell in ws[1]]
    idx = {COLMAP[h]: i for i, h in enumerate(headers) if h in COLMAP}

    with SessionLocal() as db:
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[idx["name"]] or "").strip()
            brand = str(row[idx["brand"]] or "").strip()
            if not name:
                continue

            display_name = (
                str(row[idx.get("display_name")] or "").strip()
                if "display_name" in idx
                else f"{name} | {brand}".strip(" |")
            )

            ing = (
                db.query(Ingredient)
                .filter(Ingredient.name == name, Ingredient.brand == brand)
                .first()
            )
            if not ing:
                ing = Ingredient(name=name, brand=brand)
                db.add(ing)

            ing.display_name = display_name
            ing.base_g = safe_num(row[idx.get("base_g")])
            ing.sodium_mg_100g = safe_num(row[idx.get("sodium_mg_100g")])
            ing.carbs_g_100g = safe_num(row[idx.get("carbs_g_100g")])
            ing.sugars_g_100g = safe_num(row[idx.get("sugars_g_100g")])
            ing.fiber_g_100g = safe_num(row[idx.get("fiber_g_100g")])
            ing.allulose_g_100g = safe_num(row[idx.get("allulose_g_100g")])
            ing.fat_g_100g = safe_num(row[idx.get("fat_g_100g")])
            ing.trans_fat_g_100g = safe_num(row[idx.get("trans_fat_g_100g")])
            ing.sat_fat_g_100g = safe_num(row[idx.get("sat_fat_g_100g")])
            ing.chol_mg_100g = safe_num(row[idx.get("chol_mg_100g")])
            ing.protein_g_100g = safe_num(row[idx.get("protein_g_100g")])

            memo = row[idx.get("memo")] if "memo" in idx else None
            ing.memo = str(memo).strip() if memo else None

        db.commit()

    print("✅ 원재료 DB 초기 적재 완료 (openpyxl 방식)")

if __name__ == "__main__":
    main()
